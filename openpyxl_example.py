#---------------------------------------------------------------------------------------
# Name:        openpyxl_example.py
# Purpose:     Example of using openpyxl
#
# Author:      Jesse Morgan
#
# Created:     03/03/2026
#---------------------------------------------------------------------------------------
import openpyxl
import arcpy

def read_xlsx():
    print("Reading Excel File")
    # Open the Excel file
    book = openpyxl.load_workbook(filename=r'D:\GIS\Data\example.xlsx')

    # List all worksheets
    sheets = book.sheetnames
    print(f"Worksheets: {sheets}")

    # Select a specific worksheet
    sheet = book.worksheets[0]

    # Get a cell's value
    b3 = sheet.cell(row=3, column=2)  # Or b3 = sheet["B3"]
    print(f"CELL B3: {b3.value}")

    # Read cells by iterating over the columns
    for col in sheet.iter_cols():
        for cell in col:
            print(cell.value)

    # Read cells by iterating through the rows
    for row in sheet.iter_rows():
        for cell in row:
            print(cell.value)

def read_xlsx_create_fc():
    print("Read an XLSX file and create a feature class.")
    fgdb = r"D:\GIS\Data\Data.gdb"
    in_file = r"D:\GIS\Data\example.xlsx"
    fc = 'points'
    sr = arcpy.SpatialReference(2248)
    arcpy.env.workspace = fgdb

    # Open the Excel file for reading
    book = openpyxl.load_workbook(filename=in_file)
    sheet = book.worksheets[0]

    # Create the feature class
    # arcpy.management.CreateFeatureclass(out_path=fgdb, out_name=fc, geometry_type="POINT", spatial_reference=sr)

    # Iterate through the Excel file and add points to the feature class
    with arcpy.da.InsertCursor(in_table=fc, field_names=["SHAPE@"]) as cursor:
        for i in range(2, sheet.max_row + 1):  # Skip first row which is the columns
            x = sheet.cell(row=i, column=2).value
            y = sheet.cell(row=i, column=3).value
            point = arcpy.Point(float(x), float(y))
            cursor.insertRow([point])

def main():
    # read_xlsx()
    read_xlsx_create_fc()

if __name__ == '__main__':
    main()